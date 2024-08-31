import argparse
from dataclasses import dataclass, field
import logging
import re
import string

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText

import make_kondratjev


ru_syn_pat = re.compile(r"Син.:\s(?P<lst>.*)")


def extract_ru_synonyms(ru_source: str):
    m = ru_syn_pat.search(ru_source)
    if not m:
        return []

    ##lst = re.split(r'[\, ]+', m["lst"])
    lst = []
    for s in re.split(r"[.,]+", m["lst"]):
        s = s.strip()
        if not s:
            continue
        lst.append(s)

    return lst


eo_syn_case_pat = re.compile(r"(?:\..*Sin\.:)|(?:Sin\.:)")
eo_equals_star_pat = re.compile(r"(?P<first>.*)=(?P<second>.*)\*")
eo_before_colon_pat = re.compile(r"[:,]")


def add_stripped(lst, word):
    word = word.strip()
    if not word:
        return
    lst.append(word)


def parse_sin_case(definition):
    # * if "= *" => first word + second
    m = re.search(eo_equals_star_pat, definition)
    if m:
        lst = []
        add_stripped(lst, m["first"])
        add_stripped(lst, m["second"])
        return lst

    # * cut by first ":", "," => first word
    # :TODO(ilya): case ŝarĝatoro (PIV, FE),... being cut ugly
    parts = re.split(eo_before_colon_pat, definition, maxsplit=1)
    lst = []
    if parts:
        add_stripped(lst, parts[0])

    if len(parts) <= 1:
        return lst

    # * if .split() gives one or two words => add as second
    part2 = parts[1].strip()
    if len(part2.split()) > 2:
        return lst

    add_stripped(lst, part2)
    return lst


def make_eo_syn_pattern():
    meaning_n = r"[0-9][a-z]?\.?"

    def comma_list_pat(item):
        return f"(?:{item}[,;] *)*{item}"

    meanings = comma_list_pat(meaning_n)

    source_pat = f"[A-Z,Ĉ,Ĝ,Ĥ,Ĵ,Ŝ,Ŭ]+ ?(?:{meanings})?"

    sources_pat = comma_list_pat(source_pat)

    sources_exp_pat = rf"\({sources_pat}\)?(?:[,;]|(?: =))?"

    return re.compile(sources_exp_pat)


eo_syn_sep_pat = make_eo_syn_pattern()


def parse_sin_list_case(eo_source: str):
    lst = []
    for s in re.split(eo_syn_sep_pat, eo_source):
        s = s.strip()
        if not s:
            continue
        lst.append(s)
    return lst


def parse_eo_synonyms(eo_source: str):
    parts = re.split(eo_syn_case_pat, eo_source, maxsplit=1)
    if len(parts) < 2:
        return parse_sin_list_case(eo_source)

    definition, sin_list_case = parts
    return parse_sin_case(definition) + parse_sin_list_case(sin_list_case)


def main() -> None:
    file_path = "/Users/ilya/Downloads/Attachments_lunjo@mail.ru_2024-02-06_12-50-17/Sinonimia vortaro.xlsx"  # noqa: E501

    if False:
        ##eo_source = "ekskluziva rajto (PIV 1. = prerogativo (PIV 2. "
        eo_source = "promoci = promocii*. F. promouvoir (NG). Sin.: promocii (PIV, FE, KI, BE), avansigi (KI), avancigi (BE, PIV), rangaltigi (PIV)"  # noqa: E501
        lst = parse_eo_synonyms(eo_source)
        print(lst)

    if False:
        import pandas as pd

        df = pd.read_excel(file_path)
        print(df.head())

    if False:
        ##eo_source = "perfori: tranĉeti aŭ trui laŭ vicoj, precipe al paperoj. F. perforer (NG) (BE). Sin.: trabori (PIV, KI, BE, FE), enbori (FE), trui (PIV, KI, FE), trapiki (PIV, KI), tratranĉi (KI, BR), tranĉtrui (LN), trabati (KI, BE)"  # noqa: E501
        ##eo_source = "promoci = promocii*. F. promouvoir (NG). Sin.: promocii (PIV, FE, KI, BE), avansigi (KI), avancigi (BE, PIV), rangaltigi (PIV)"  # noqa: E501
        ##eo_source = "asalti: kurataki. F. donne l' assaut (NG). Sin.: ataki (KI, BV, PIV), sturmi (BV, PIV), ŝturmi (KI), atakegi (KI), breĉataki (KI), ŝtormataki (FE), saltataki (KI), kurpenetri (LN)"  # noqa: E501
        # eo_source = "kloto (teks.): speco de kotona tolo, uzata precipe por antaŭtukoj, labor-vestoj kaj kiel subŝtofoj; kalikoto*. G. Cloth; F. calicot (NG). Sin.: kalikoto (PIV, FE, BE, KI)"  # noqa: E501
        ##eo_source = "ŝarĝatoro (PIV, FE), levĉaro (PIV, FE), ŝovelmaŝino (PIV, FE, BE), ĉarelo (PIV), altleva ĉarelo (PIV), flankforka ĉarelo (PIV), ekskavatoro: maŝino por fosi grundon kaj forigi la elskrapitan materialon. F. excavateur (NG) (KI, BE). Sin.: elkavatoro (PIV, FE), elkavigilo (FE), elkavigmaŝino (FE), terfosmaŝino (KI), fosmaŝino (PIV),  ŝovelmaŝino (PIV, FE, BE), skrapmaŝino (PIV)"  # noqa: E501
        eo_source = "cidi: mortigi, neniigi, ekstermi: gentocido, herbocidilo, insektocidilo. F. tier, anéantir, exterminer (NG). Sin.: ekstermi (PIV, FE, KI, BV), neniigi (KI, PIV,FE, BV), malaperigi (KI, PIV), mortigi (KI, BV, PIV), pereigi (PIV, KI, BV)"  # noqa: E501
        ##eo_source = "aserti sin: agi aŭ sin esprimi kun forto antaŭ alia(j) persono(j), ne lasante sin superi de iu aŭtoritata, nek la eventuala kontraŭa premo de grupo. F. s' affirmer (NG). Sin.: manifestiĝi (FE), firmiĝi (FE)"  # noqa: E501

        ##eo_source = "aberacio (PIV 1."

        # * cut by first dot, if exists
        # :TODO(ilya): case kloto (teks.): fails it
        parts = re.split(eo_syn_case_pat, eo_source, maxsplit=1)
        assert len(parts) >= 2

        definition = parts[0]

        lst = parse_sin_case(definition)
        print(lst)

    if False:
        ##eo_source = "apologio (BE), laŭdego (KI, W, BV), panegiro (PIV, W, FE, BE), panegiraĵo (KI), superlaŭdo (BR), glorado (PIV, KI, W, FE, BE), apoteozo (PIV 2."  # noqa: E501
        eo_source = """najlotirilo (BE), tenajlo (PIV, BE),  plattenajlo (K), platbeka tenajl(et)o (K);
alĝustigebla tenajlo (K), etendebla tenajlo (K), universala tenajlo (K), kombinita tenajlo (K), tenajlo por diversaj uzo (K); pinĉilo (PIV), prenilo (PIV, FE), pinĉilo (PIV, FE, BE, pinĉilego (PIV), bekpinĉilo (PIV, K)
"""  # noqa: E501
        ##eo_source = "aberacio (PIV 1."
        ##eo_source = "septo (PIV 2,3, BE 1., vando (PIV, KI, BE), barilo (KI), mureto (K), fakmuro (K),  ŝirmilo (KI), paravento (KI) = ventŝirmilo (PIV, FE), ekranego (PIV, V), ventŝirmilo (PIV), kaŝilo (BR), faldebla ŝirmilo (V, T)"  # noqa: E501
        ##eo_source = "ekskluziva rajto (PIV 1. = prerogativo (PIV 2. "
        eo_source = " aĥ!  (PIV), ho! (PIV, KI,  BE), he! (KI), oj! (PIV, BE), ha! (PIV, KI),  eheu!  (NG), ve! (PIV 1, FE, KI,  BV), bedaŭrinde! (PIV, KI, BV, FE),  ho ve! (PIV, BR), malplezure! (BE)"  # noqa: E501
        eo_source = "sed (PIV, KI, BE, ŜR), tamen (PIV, BE)"

        lst = parse_sin_list_case(eo_source)
        print(lst)

    if False:
        ru_source = "адепт. Син.: апологет,  защитник, ревнитель, заступник, поклонник, поборник, борец,  (фанатичный) приверженец, панегирист, зелот, сторонник, энтузиаст, последователь, пропонент"  # noqa: E501
        print(extract_ru_synonyms(ru_source))

    if True:
        ##argv = None # sys.argv
        argv = [
            ##"--gen_source_only",
            file_path,
            "/Users/ilya/opt/programming/eoru/tmp/novikova/sinonimoj.txt",
        ]

        logging.basicConfig(level=logging.INFO)

        parser = argparse.ArgumentParser()
        parser.add_argument(
            "--gen_source_only",
            action="store_true",
            help="generate source texts for stardict compiler only",
        )
        parser.add_argument("src_fname")
        parser.add_argument("dst_fname")
        args = parser.parse_args(argv)

        engine_kwargs = {
            "read_only": True,
            "data_only": True,
            "keep_links": False,
            "rich_text": True,
        }

        wb = load_workbook(
            args.src_fname,
            **engine_kwargs,
        )

        class Column:
            RU = 2
            EO = 3
            EXAMPLES = 4
            REMARKS = 5

        @dataclass
        class ArticleSource:
            ru: list = field(default_factory=list)
            eo: list = field(default_factory=list)
            examples: list = field(default_factory=list)
            remarks: list = field(default_factory=list)

        @dataclass
        class Article:
            src: ArticleSource
            name: str
            ru_synonyms: list[str]
            eo_synonyms: list[str]

        articles: list[Article] = []
        for sheet in wb.worksheets:
            logging.info(f"processing {sheet.title}...")
            first_article_index = len(articles)

            if sheet.title == "Й":
                logging.warning("letter Й is skipped due to different formatting, TODO")
                continue

            content_shift = 0
            if sheet.title in ["В", "Е"]:
                content_shift = 1

            class state:
                cur_article = ArticleSource()

            def add_article():
                article = state.cur_article
                state.cur_article = ArticleSource()

                if not article.ru:
                    return
                assert article.eo

                root_cell = article.ru[0]
                value = root_cell.value
                if not value:
                    if value is not None:
                        logging.warning(
                            f"root cell {root_cell.coordinate} has empty name"
                        )

                    for cell in article.ru:
                        if cell.value:
                            logging.warning(
                                f"root cell {root_cell.coordinate} has empty name and not empty content {cell.value}"
                            )
                            break

                    return

                if isinstance(value, CellRichText):
                    name = str(value[0])
                elif isinstance(value, str):
                    name = value
                else:
                    name = str(value)
                    logging.warning(
                        f"unknown cell content type: {root_cell.coordinate}, {name}"
                    )

                if not name:
                    logging.warning(
                        f"root cell {root_cell.coordinate} has empty name as string"
                    )
                    return

                # велоспорт. Син.: велосипедный спорт, ...
                name = name.split(".", maxsplit=1)[0]
                # абазги,
                name = name.strip(",." + string.whitespace)

                if not name:
                    logging.warning(
                        f"root cell {root_cell.coordinate} has whitespaces in name only"
                    )
                    return

                # *
                ru_synonyms = extract_ru_synonyms(str(value))

                # *
                eo_synonyms = []
                for item in article.eo:
                    value = item.value
                    if not value:
                        continue
                    res = parse_eo_synonyms(str(value))
                    eo_synonyms.extend(res)
                ##eo_synonyms = list(set(eo_synonyms))

                articles.append(
                    Article(
                        src=article,
                        name=name,
                        ru_synonyms=ru_synonyms,
                        eo_synonyms=eo_synonyms,
                    )
                )

            for row in sheet.iter_rows(min_row=2):

                article_end = False
                for column, cell in enumerate(row, start=1):
                    content_cln = column - content_shift
                    if content_cln == Column.RU:
                        top_style = cell.border.top.style

                        if top_style:
                            if top_style != "medium":
                                logging.warning(
                                    f"unknown border style for article: {top_style}, {cell.coordinate}"
                                )

                            add_article()

                        state.cur_article.ru.append(cell)

                        bottom_style = cell.border.bottom.style
                        if bottom_style:
                            if bottom_style != "medium":
                                logging.warning(
                                    f"unknown border style for article: {bottom_style}, {cell.coordinate}"
                                )

                            article_end = True

                    elif content_cln == Column.EO:
                        state.cur_article.eo.append(cell)
                    elif content_cln == Column.EXAMPLES:
                        state.cur_article.examples.append(cell)
                    elif content_cln == Column.REMARKS:
                        state.cur_article.remarks.append(cell)

                if article_end:
                    add_article()

            add_article()

            if first_article_index == len(articles):
                logging.warning("no articles added for the sheet")
            else:
                logging.info(
                    f"first article: {articles[first_article_index].src.ru[0].value}"
                )
                logging.info(f"last article: {articles[-1].src.ru[0].value}")

        logging.info(f"total: {len(articles)}")

        def append_cell_content(lst, value):
            if value:
                lst.append(str(value).strip())

        with make_kondratjev.dictionary_generator(
            args.dst_fname,
            gen_source_only=args.gen_source_only,
        ) as on_parsed_article:
            for article in articles:
                name = article.name
                synonyms = set(article.ru_synonyms + article.eo_synonyms)
                synonyms.discard(name)

                all_names = [name] + list(synonyms)

                bad_name = False
                for name in all_names:
                    # muslin||o текс. мусли́н; шифо́н K)
                    #      kotona ~o хлопчатобума́жный мусли́н; кисея́ (K)
                    # ~a мусли́новый; шифо́новый; кисе́йный (K)
                    for c in ["|", "\n"]:
                        if c in name:
                            logging.warning(f"synonym {name}, bad character occurrence: {repr(c)}")
                            bad_name = True
                            break
                if bad_name:
                    continue

                article_blocks: list[str] = []
                src = article.src
                for i in range(len(src.ru)):
                    first_lst: list[str] = []
                    append_cell_content(first_lst, src.ru[i].value)
                    append_cell_content(first_lst, src.eo[i].value)

                    second_lst: list[str] = []
                    if src.examples[i].value:
                        append_cell_content(second_lst, "Примеры употребления:")
                        append_cell_content(second_lst, src.examples[i].value)
                    if src.remarks[i].value:
                        append_cell_content(second_lst, "Примечания:")
                        append_cell_content(second_lst, src.remarks[i].value)

                    block = ["\n".join(first_lst)]
                    if first_lst and second_lst:
                        block.append("\n\n")
                    block.append("\n".join(second_lst))

                    article_blocks.append("".join(block))

                article_txt = "\n\n".join(article_blocks)

                on_parsed_article(all_names, article_txt)

    if False:
        engine_kwargs = {
            "read_only": True,
            "data_only": True,
            "keep_links": False,
            "rich_text": True,
        }

        wb = load_workbook(
            file_path,
            **engine_kwargs,
        )

        ##pprint.pprint([ws.title for ws in wb.worksheets])

        # for sheet in wb.worksheets:
        #    print(sheet.title)

        sheet = None  # type: ignore
        for sh in wb.worksheets:
            if sh.title == "А":
                sheet = sh
                break

        assert sheet
        ##print(sheet)

        # article start vs. not
        b2 = None
        b3 = None
        # eo keys
        c270 = None

        for row in sheet.iter_rows():
            ##print(row)
            for cell in row:
                if cell.data_type == "n":
                    # EmptyCell
                    continue

                if cell.coordinate == "B2":
                    b2 = cell
                if cell.coordinate == "B3":
                    b3 = cell
                if cell.coordinate == "C270":
                    c270 = cell

                print(f"{cell.coordinate}={cell.value}", end="\t")

            print()

        print(b2, b3, c270)


if __name__ == "__main__":
    main()
